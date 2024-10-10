import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo

from unidecode import unidecode
import re
import os

#df_emissoes = pd.read_excel(r'C:\Users\Thomas\Downloads\Base_v2\Emissao\Dados_Emissao\Agrupado\Emissoes_agrupado.xlsx')

#df_renovacoes = pd.read_excel(r'C:\Users\Thomas\Downloads\Base_v2\Renovacao\Dados_Renovacao\Agrupado\Renovacoes_agrupado.xlsx')

class ComparativoDaily():
    """Takes the path of renovações agrupados e emissoes agrupado"""

    def __init__(self, emissoes, renovacoes) -> None:
        self.emissoes = emissoes
        self.renovacoes = renovacoes
        self.foco = r'C:\Users\Thomas\Downloads\Base_v2\zOutros\corretoresfoco.xlsx'
        #Call Func
        self.read_excel()
        #Grades
        self.df_grades = self.grades()
        #Pivot inspetor
        self.pivot_emissao_insp = self.pivot(self.emissoes, 1)
        self.pivot_renovacao_insp = self.pivot(self.renovacoes, 1)
        #Pivot Corretor
        self.pivot_emissao_corr = self.pivot(self.emissoes, 2)
        self.pivot_renovacao_corr = self.pivot(self.renovacoes, 2)
        

        #self.segmentados()


        self.pla_separadas(self.tabela_promocao(),r'C:\Users\Thomas\Downloads\base_relatorios\campanha_mais_cinco',['Realizado %'])
        self.pla_separadas(self.df_grades, r'C:\Users\Thomas\Downloads\base_relatorios\planilhas_foco')
        #self.daily_table()
        self.add_foco(r'C:\Users\Thomas\Downloads\base_relatorios\planilhas_foco')
    
    def read_excel(self):
        """Read the path as excel"""
        self.emissoes = pd.read_excel(self.emissoes, index_col=None)
        self.renovacoes = pd.read_excel(self.renovacoes, index_col=None)
        self.foco = pd.read_excel(self.foco, index_col=None)
        self.seg_residencial = self.foco['Residencial']
        self.seg_demais_re = self.foco['Demais RE']
    
    def segmentados(self):
        #self.seg_residencial = pd.merge(self.foco[['Residencial']], self.df_grades, how = 'inner', on = 'Corretor')
        self.foco = [unidecode(col).replace('ç', 'c') for col in self.foco]
        #resultado = pd.merge(self.foco[['Residencial']], self.df_grades, how='inner', left_on='Residencial', right_on='Corretor')[['Corretor', 'Inspetor de producao']]
        print(self.foco)
        
    def pivot(self, df, op):
        """
        [1]Pivot = inspetor de producao (Index). Produtos Coluna
        [2]Pivot = Correotr (Index). Produtos Coluna
        """
        df['Ramo Seguro'].fillna('Zerado', inplace=True)
        
        if op == 1:
            df = df.pivot_table(index='Inspetor de producao', columns='Ramo Seguro', aggfunc='size', fill_value=0)
        if op == 2:
            df = df.pivot_table(index='Corretor', columns='Ramo Seguro', aggfunc='size', fill_value=0)
            df = pd.merge(df, self.df_grades, how = 'inner', on = 'Corretor')
        
        df['Demais RE'] = df[['Condominio', 'Empresarial', 'Equipamento', 'Equipamento Agricola']].sum(axis=1)  

        return df

    def grades(self):
        """Pega a lista de grades (Corretor/Comercial), mais recente (com base nas emissoes agrupado)"""
        return self.emissoes.drop_duplicates(subset='Corretor')[['Corretor','Inspetor de producao']]
    
    
    def list_files_in_folder(self, folder_path):
        #Generico
        files = []
        for file in os.listdir(folder_path):
            if os.path.isfile(os.path.join(folder_path, file)) and not file.startswith('~'):
                files.append(os.path.join(folder_path, file))
        return files


    def criar_aba_com_dataframe(self, nome_arquivo, nome_aba, dataframe):
        #Generico
        # Carrega o workbook existente
        wb = load_workbook(nome_arquivo)
        
        # Cria uma nova aba
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            writer.book = wb
            dataframe.to_excel(writer, sheet_name=nome_aba, index=False)
        
        # Salva as alterações no arquivo
        wb.save(nome_arquivo)
    
    def add_foco(self, pp_folder):
        #pp_folder = princiapal_path_folder
        lista_planilhas = self.list_files_in_folder(pp_folder)
        colunas = {
            'Foco Residencial': [],
            'Foco Demais RE': []
        }

        # Criando o DataFrame vazio com as colunas nomeadas
        df = pd.DataFrame(colunas)
        for i in lista_planilhas:
            self.criar_aba_com_dataframe(i, 'foco', df)

    
    def tabela_promocao(self):

        df_renovacao = self.pivot_renovacao_corr
        df_renovacao['Auto Target'] = df_renovacao['Auto'] + 5
        df_renovacao['Auto Renovacao'] = df_renovacao['Auto']
        df_renovacao = df_renovacao[['Corretor', 'Auto Renovacao', 'Auto Target', 'Inspetor de producao']]

        df_emitido = self.pivot_emissao_corr
        df_emitido['Auto Emitido'] = df_emitido['Auto']
        df_emitido = df_emitido[['Corretor', 'Auto Emitido']]

        df1 = pd.merge(df_renovacao, df_emitido, how='inner', on='Corretor')
        df1['Realizado %'] = df1['Auto Emitido'] / df1['Auto Target']
        df1 = df1[['Corretor', 'Auto Renovacao', 'Auto Emitido', 'Auto Target', 'Realizado %', 'Inspetor de producao']]
        #df1 = df1[df1['Auto Target'] >= 10]
        df1 = df1[(df1['Auto Target'] >= 10) | (df1['Corretor'] == 'FR SJC CORRETORA DE SEGUROS LTDA') | (df1['Corretor'] == 'BETEL VALE CORR DE SEGS LTDA')]
        df1.to_excel(r'C:\Users\Thomas\Downloads\base_relatorios\campanha_mais_cinco.xlsx', index=False)
        return df1

    def pla_separadas(self, df, path, lista=None):
        """Takes a group of data and separetes by commercial"""
        unique_ins = df['Inspetor de producao'].unique().tolist()
        l_plans = []
        for i in unique_ins:
            df1 = df[df['Inspetor de producao'] == i]
            
            self.export_to_excel_with_format(df1, self.join_path(path, i, '.xlsx'),lista)
    
    def join_path(self, path, file, type):
        nome = unidecode(file)
        nome = re.sub(r'[\\/]+', '', nome)
        caminho_final_img = os.path.join(path, nome + type)
        return caminho_final_img

    
    def export_to_excel_with_format(self, df, file_path, percentage_columns):
        """
        Exporta um DataFrame para um arquivo Excel com colunas formatadas como porcentagem.
        O cabeçalho de todas as colunas será formatado com fundo vermelho, letras brancas e título em negrito.
        Adiciona um filtro à planilha.
        
        Parâmetros:
        df (pd.DataFrame): O DataFrame a ser exportado.
        file_path (str): O caminho do arquivo Excel a ser criado.
        percentage_columns (list): Lista com os nomes das colunas a serem formatadas como porcentagem.
        """
        # Exporta o DataFrame para um arquivo Excel
        df.to_excel(file_path, index=False)
        
        # Carrega o arquivo Excel
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Cria um estilo para formatação de porcentagem
        percentage_style = NamedStyle(name='percentage')
        percentage_style.number_format = numbers.FORMAT_PERCENTAGE_00
        
        # Cria um estilo para o cabeçalho
        header_style = NamedStyle(name='header')
        header_style.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        header_style.font = Font(color='FFFFFF', bold=True)  # Título em negrito
        
        # Aplica a formatação de porcentagem às colunas especificadas
        if percentage_columns:
            for col_name in percentage_columns:
                col_idx = df.columns.get_loc(col_name) + 1  # Obtém o índice da coluna (1-base)
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                    for c in cell:
                        c.style = percentage_style
        
        # Aplica a formatação ao cabeçalho de todas as colunas
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=1, column=col_idx).style = header_style
        
        # Adiciona um filtro à planilha
        ws.auto_filter.ref = ws.dimensions  # Define o autofiltro para todas as células preenchidas
        
        # Salva o arquivo Excel com a formatação aplicada
        wb.save(file_path)


    def daily_table(self):
        df_emissao = self.pivot_emissao_insp[['Auto', 'Residencial Sob Medida', 'Demais RE']]
        df_emissao = df_emissao.rename(columns=lambda x: x + ' 2024')

        df_renovacao = self.pivot_renovacao_insp[['Auto', 'Residencial Sob Medida', 'Demais RE']]
        df_renovacao = df_renovacao.rename(columns=lambda x: x + ' 2023')

        df_renovacao.rename(index={'Jorge Henrique de Souza Vasconcellos': 'Mateus Rodrigues da Silva'},inplace=True)
        
        df_merged = pd.merge(df_emissao, df_renovacao, how='right', on='Inspetor de producao')

        df_merged['Auto %'] = df_merged['Auto 2024'] / df_merged['Auto 2023']
        df_merged['Residencial %'] = df_merged['Residencial Sob Medida 2024'] / df_merged['Residencial Sob Medida 2023']
        df_merged['Demais RE %'] = df_merged['Demais RE 2024'] / df_merged['Demais RE 2023']

        df_merged['Residencial 2023'] = df_merged['Residencial Sob Medida 2023']
        df_merged['Residencial 2024'] = df_merged['Residencial Sob Medida 2024']

        df_merged = df_merged[['Auto 2023', 'Auto 2024', 'Auto %', 'Residencial 2023', 'Residencial 2024', 'Residencial %', 'Demais RE 2023', 'Demais RE 2024', 'Demais RE %']]

        print(df_merged)
    


ComparativoDaily(r'C:\Users\Thomas\Downloads\Base_v2\Emissao\Dados_Emissao\Agrupado\Emissoes_agrupado.xlsx', r'C:\Users\Thomas\Downloads\Base_v2\Renovacao\Dados_Renovacao\Agrupado\Renovacoes_agrupado.xlsx')
